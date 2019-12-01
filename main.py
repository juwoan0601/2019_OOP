from openpyxl import load_workbook
from collections import OrderedDict
from sim import *
from database import *
from pymongo import MongoClient
from bson import json_util
import json

##USEFUL INFORMATION ABOOUT DATA##
# database의 data형식>>>
# data는 json형식으로 name, unchange, change, restaurant로 이루어져 있음
# name: 음식점 이름(string), unchange: 맛, 주재료 등 음식 고유의 property(string --> code내에서는 binary로 바꿔서 작동, 관련 함수 database에서 제공)
# change: 날씨, 계절 등의 정보 저장(string -> code 내에서는 int array로 바꿔서 작동, 관련함수 database에서 제공), restaurant: 음식점 정보(string)
# name은 띄어쓰기가 없음

# Data example>>
# ---------------------------------------------------------
# |name         unchange        change          restaurant|
# |--------------------------------------------------------
# |두루치기정식   101011...     12/13/99/1.../    인천식당  |
# |...           ...           ...               ...      |

# Data 호출방법>>
# json객체 a의 name 호출: a['name'] 

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>여기서부터 추가된 코드
#database에서 데이터를 받아들이고 이를 json형식의 list로 만든 다음 dictionary 형식으로 바꾼다
dic1 = {} #dic1은 name <==> unchange를 연결
dic2 = {} #dic2는 name <==> change를 연결
dic3 = {} #dic3는 name <==> restaurant를 연결
l=FoodList(collection) #l은 json형식의 전체 음식 리스트, 리스트의 한 요소에는 name, unchange(맛, 주재료)
for i in l:
    dic1[i['name']] = ReturnBin(collection, i['name'], unchange) #database에서 unchange, chage 모두 string이므로 각각 binary vector와  int vector로 바꿔준다. 
    dic2[i['name']] = ReturnBin(collection, i['name'], change)
    dic3[i['name']] = i['restaurant']


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>여기서부터는 예전 코드
# 엑셀 파일로부터 데이터를 받아들임
#def encode(excel):
#   wb = load_workbook(excel, data_only=True)
#    ws = wb["Sheet1"]

#    dic = {ws.cell(2, 1).value: 0}
#    for i in range(1, ws.max_column):
#        if ws.cell(2, i + 1).value:
#            dic[ws.cell(2, 1).value] |= (1 << i)

#   for i in range(2, ws.max_row):
#        for j in range(1, ws.max_column):
#            if ws.cell(i + 1, j + 1).value:
#                dic.setdefault(ws.cell(i + 1, 1).value, 0)
#                dic[ws.cell(i + 1, 1).value] |= (1 << j)

# 딕셔너리에 저장된 key값과 인코딩된 비트의 속성을 출력
#    for recipe in dic:
#        print()
#        print(recipe)
#        for i in range(1, ws.max_column):
#            if dic[recipe] & (1 << i):
#                print("# " + ws.cell(1, i+1).value)
#    return dic

# encode(input("input .xlsx data : "))
# dictionary는 순서가 없기때문에 순서를 만들어줌.
# 예시 data
# f = {"치킨":542, "피자":436, "탕수육":204, "짜장면":242, "짬뽕": 231}

while True:
    f = encode(input("input .xlsx data : "))
    mode = input("mode (d: Dot product, h: Harmonic mean, c: Cosine) : ")

    if(mode != 'd'and mode !='h'and mode != 'c'):
        print("mode를 잘못입력함.")
        break

    barcode = input("비교하고자 하는 대상 바코드 입력: ")
    # 입력 예시 : 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 1, 0
    barcode = barcode.replace(", ", "")
    barcode = int(barcode,2)
    barcode = IntToBinV(barcode)

    food = OrderedDict(sorted(f.items()))
    print("Similarity Map\n")


    # 정렬되지 않은 Similarity Map
    #Sim_not_sort.Result_print(mode, barcode, food)
    # 정렬된 Similarity Map
    Sim_sort.Result_print(mode, barcode, food)
